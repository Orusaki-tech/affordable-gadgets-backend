import csv

from django.core.management.base import BaseCommand, CommandError

try:
    import openpyxl

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class Command(BaseCommand):
    help = "Export a specific Excel sheet to CSV"

    def add_arguments(self, parser):
        parser.add_argument("excel_path", type=str, help="Path to Excel file")
        parser.add_argument(
            "--sheet",
            type=str,
            default="All Stock Buy Backs & Imports",
            help='Sheet name to export (default: "All Stock Buy Backs & Imports")',
        )
        parser.add_argument(
            "--output",
            type=str,
            help="Output CSV path (default: excel path with .csv extension)",
        )

    def handle(self, *args, **options):
        excel_path = options["excel_path"]
        sheet_name = options["sheet"]

        if not HAS_OPENPYXL:
            raise CommandError("openpyxl is required. Install with: pip install openpyxl")

        try:
            # Try loading with data_only=True to get calculated formula values
            # This only works if Excel has calculated the formulas before saving
            try:
                workbook = openpyxl.load_workbook(excel_path, data_only=True)
            except Exception:
                # Fall back to normal loading if data_only fails
                workbook = openpyxl.load_workbook(excel_path)
        except FileNotFoundError:
            raise CommandError(f"Excel file not found: {excel_path}")
        except Exception as exc:
            raise CommandError(f"Failed to load Excel: {exc}")

        if sheet_name not in workbook.sheetnames:
            raise CommandError(
                f'Sheet "{sheet_name}" not found. Available: {", ".join(workbook.sheetnames)}'
            )

        sheet = workbook[sheet_name]

        csv_path = options["output"] or excel_path.replace(".xlsx", ".csv").replace(".xls", ".csv")

        with open(csv_path, "w", newline="", encoding="utf-8") as csvfile:
            writer = csv.writer(csvfile)
            for row in sheet.iter_rows(values_only=True):
                cleaned_row = ["" if cell is None else str(cell) for cell in row]
                writer.writerow(cleaned_row)

        self.stdout.write(self.style.SUCCESS(f'Exported sheet "{sheet_name}" to {csv_path}'))

